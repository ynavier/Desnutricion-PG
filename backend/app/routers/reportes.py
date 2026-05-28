"""
Router de reportes.
POST /reportes/generar    — genera y cachea un reporte
GET  /reportes/historial  — últimos 20 reportes generados
GET  /reportes/{id}       — reporte completo desde caché
GET  /reportes/{id}/excel — descarga XLSX
"""
import io
import uuid
from datetime import datetime
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from app.auth.dependencies import get_current_user
from app.services.reportes_data import (
    reporte_nutricional,
    reporte_epidemiologico,
    reporte_pacientes,
    reporte_modelo,
)
from app.services.ai_reportes import analizar_reporte

router = APIRouter(prefix='/reportes', tags=['reportes'])

# ── Almacenamiento en memoria ─────────────────────────────────────────────────
_historial: list[dict] = []          # metadatos de los últimos 20 reportes
_cache_reportes: dict[str, dict] = {} # id → reporte completo


# ── Schemas ───────────────────────────────────────────────────────────────────

class GenerarRequest(BaseModel):
    tipo:           str                    # nutricional | epidemiologico | pacientes | modelo
    fecha_desde:    Optional[str] = None
    fecha_hasta:    Optional[str] = None
    zona:           Optional[str] = None
    establecimiento: Optional[str] = None


# ── Helpers ───────────────────────────────────────────────────────────────────

_TITULOS = {
    'nutricional':    'Reporte Nutricional',
    'epidemiologico': 'Reporte Epidemiológico',
    'pacientes':      'Listado de Pacientes',
    'modelo':         'Desempeño del Modelo ML',
}

_DATA_FN = {
    'nutricional':    reporte_nutricional,
    'epidemiologico': reporte_epidemiologico,
    'pacientes':      reporte_pacientes,
    'modelo':         reporte_modelo,
}


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post('/generar')
async def generar_reporte(
    body: GenerarRequest,
    user: dict = Depends(get_current_user),
):
    tipo = body.tipo.lower()
    if tipo not in _DATA_FN:
        raise HTTPException(status_code=400, detail=f'Tipo de reporte inválido: {tipo}')

    filtros = {
        'fecha_desde':    body.fecha_desde,
        'fecha_hasta':    body.fecha_hasta,
        'zona':           body.zona,
        'establecimiento': body.establecimiento,
    }

    # 1. Obtener datos
    data_fn = _DATA_FN[tipo]
    datos = await data_fn(filtros)

    # 2. Análisis IA
    analisis = await analizar_reporte(
        tipo=tipo,
        kpis=datos.get('kpis', {}),
        distribucion=datos.get('distribucion'),
    )

    # 3. Construir reporte
    reporte_id = str(uuid.uuid4())
    ahora = datetime.now().isoformat(timespec='seconds')

    reporte = {
        'id':     reporte_id,
        'tipo':   tipo,
        'titulo': _TITULOS.get(tipo, tipo),
        'filtros': filtros,
        'generado': ahora,
        'kpis':     datos.get('kpis', {}),
        'charts':   datos.get('charts', {}),
        'tabla':    datos.get('tabla', []),
        'analisis': analisis,
    }

    # 4. Guardar en caché e historial
    _cache_reportes[reporte_id] = reporte

    meta = {
        'id':       reporte_id,
        'tipo':     tipo,
        'titulo':   _TITULOS.get(tipo, tipo),
        'generado': ahora,
        'fecha_desde': body.fecha_desde,
        'fecha_hasta': body.fecha_hasta,
        'fuente_ia':   analisis.get('fuente', 'estatica'),
    }
    _historial.insert(0, meta)
    # Mantener solo los últimos 20
    del _historial[20:]

    return reporte


@router.get('/historial')
async def obtener_historial(user: dict = Depends(get_current_user)):
    return _historial


@router.get('/{reporte_id}')
async def obtener_reporte(
    reporte_id: str,
    user: dict = Depends(get_current_user),
):
    reporte = _cache_reportes.get(reporte_id)
    if not reporte:
        raise HTTPException(status_code=404, detail='Reporte no encontrado o expirado')
    return reporte


@router.get('/{reporte_id}/excel')
async def descargar_excel(
    reporte_id: str,
    user: dict = Depends(get_current_user),
):
    reporte = _cache_reportes.get(reporte_id)
    if not reporte:
        raise HTTPException(status_code=404, detail='Reporte no encontrado o expirado')

    xlsx_bytes = _generar_excel(reporte)
    titulo_safe = reporte['titulo'].replace(' ', '_').replace('/', '_')
    filename = f"{titulo_safe}_{reporte['generado'][:10]}.xlsx"

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'},
    )


# ── Generación de Excel ───────────────────────────────────────────────────────

def _generar_excel(reporte: dict) -> bytes:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError as e:
        raise HTTPException(status_code=500, detail=f'openpyxl no instalado: {e}')

    HDR_FILL = PatternFill('solid', fgColor='4FB4D2')
    HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
    TITLE_FONT = Font(bold=True, size=14, color='1A1F2B')
    LABEL_FONT = Font(bold=True, size=10, color='54606E')
    CENTER = Alignment(horizontal='center', vertical='center')

    wb = openpyxl.Workbook()

    # ── Hoja 1: Resumen ────────────────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = 'Resumen'

    ws1.column_dimensions['A'].width = 28
    ws1.column_dimensions['B'].width = 32
    ws1.column_dimensions['C'].width = 20
    ws1.column_dimensions['D'].width = 20

    row = 1
    ws1.merge_cells(f'A{row}:D{row}')
    cell = ws1.cell(row=row, column=1, value=reporte['titulo'])
    cell.font = TITLE_FONT
    cell.alignment = CENTER

    row += 1
    filtros = reporte.get('filtros', {})
    periodo = f"{filtros.get('fecha_desde', 'N/A')} — {filtros.get('fecha_hasta', 'N/A')}"
    ws1.merge_cells(f'A{row}:D{row}')
    cell = ws1.cell(row=row, column=1, value=f'Período: {periodo}')
    cell.font = Font(size=10, color='54606E')
    cell.alignment = CENTER

    row += 1
    ws1.merge_cells(f'A{row}:D{row}')
    cell = ws1.cell(row=row, column=1, value=f'Generado: {reporte.get("generado", "")}')
    cell.font = Font(size=10, color='54606E')
    cell.alignment = CENTER

    row += 2
    ws1.cell(row=row, column=1, value='INDICADORES CLAVE').font = Font(bold=True, size=11)
    row += 1

    ws1.cell(row=row, column=1, value='Indicador').font = HDR_FONT
    ws1.cell(row=row, column=1).fill = HDR_FILL
    ws1.cell(row=row, column=2, value='Valor').font = HDR_FONT
    ws1.cell(row=row, column=2).fill = HDR_FILL
    row += 1

    kpis = reporte.get('kpis', {})
    for k, v in kpis.items():
        label = k.replace('_', ' ').title()
        ws1.cell(row=row, column=1, value=label).font = LABEL_FONT
        ws1.cell(row=row, column=2, value=v)
        row += 1

    # Análisis IA
    analisis = reporte.get('analisis', {})
    if analisis:
        row += 1
        ws1.cell(row=row, column=1, value='ANÁLISIS').font = Font(bold=True, size=11)
        row += 1

        resumen = analisis.get('resumen', '')
        if resumen:
            ws1.cell(row=row, column=1, value='Resumen').font = LABEL_FONT
            ws1.merge_cells(f'B{row}:D{row}')
            ws1.cell(row=row, column=2, value=resumen).alignment = Alignment(wrap_text=True)
            ws1.row_dimensions[row].height = 50
            row += 1

        hallazgos = analisis.get('hallazgos', [])
        if hallazgos:
            ws1.cell(row=row, column=1, value='Hallazgos').font = LABEL_FONT
            row += 1
            for h in hallazgos:
                ws1.merge_cells(f'B{row}:D{row}')
                ws1.cell(row=row, column=2, value=f'• {h}').alignment = Alignment(wrap_text=True)
                ws1.row_dimensions[row].height = 35
                row += 1

        recomendaciones = analisis.get('recomendaciones', [])
        if recomendaciones:
            ws1.cell(row=row, column=1, value='Recomendaciones').font = LABEL_FONT
            row += 1
            for r in recomendaciones:
                ws1.merge_cells(f'B{row}:D{row}')
                ws1.cell(row=row, column=2, value=f'• {r}').alignment = Alignment(wrap_text=True)
                ws1.row_dimensions[row].height = 35
                row += 1

    # ── Hoja 2: Distribución ──────────────────────────────────────────────────
    ws2 = wb.create_sheet('Distribución')
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 16

    charts = reporte.get('charts', {})
    dist_estado = charts.get('distribucion_estado', [])

    ws2.cell(row=1, column=1, value='Estado Nutricional').font = HDR_FONT
    ws2.cell(row=1, column=1).fill = HDR_FILL
    ws2.cell(row=1, column=2, value='Cantidad').font = HDR_FONT
    ws2.cell(row=1, column=2).fill = HDR_FILL

    for i, item in enumerate(dist_estado, start=2):
        ws2.cell(row=i, column=1, value=item.get('estado', ''))
        ws2.cell(row=i, column=2, value=item.get('cantidad', 0))

    # ── Hoja 3: Tendencia ─────────────────────────────────────────────────────
    tendencia = charts.get('tendencia_mensual') or charts.get('tendencia_casos') or charts.get('evolucion_accuracy')
    if tendencia:
        ws3 = wb.create_sheet('Tendencia')
        ws3.column_dimensions['A'].width = 16

        if tendencia:
            # Encabezados dinámicos
            first = tendencia[0]
            cols = list(first.keys())
            for ci, col in enumerate(cols, start=1):
                cell = ws3.cell(row=1, column=ci, value=col.replace('_', ' ').title())
                cell.font = HDR_FONT
                cell.fill = HDR_FILL
                ws3.column_dimensions[
                    openpyxl.utils.get_column_letter(ci)
                ].width = 18

            for ri, item in enumerate(tendencia, start=2):
                for ci, col in enumerate(cols, start=1):
                    ws3.cell(row=ri, column=ci, value=item.get(col, ''))

    # ── Hoja 4: Pacientes (solo para reporte tipo pacientes) ──────────────────
    tabla = reporte.get('tabla', [])
    if tabla:
        ws4 = wb.create_sheet('Pacientes')
        headers = ['ID', 'Nombre', 'Edad', 'Sexo', 'Municipio', 'Establecimiento',
                   'Último Estado', 'Último Control', 'N° Controles', 'Seguimiento']
        keys    = ['id', 'nombre', 'edad', 'sexo', 'municipio', 'establecimiento',
                   'ultimo_estado', 'ultimo_control', 'n_controles', 'seguimiento']

        for ci, h in enumerate(headers, start=1):
            cell = ws4.cell(row=1, column=ci, value=h)
            cell.font = HDR_FONT
            cell.fill = HDR_FILL
            ws4.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = 18

        for ri, row_data in enumerate(tabla[:200], start=2):
            for ci, key in enumerate(keys, start=1):
                val = row_data.get(key, '')
                if isinstance(val, bool):
                    val = 'Sí' if val else 'No'
                ws4.cell(row=ri, column=ci, value=val)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
